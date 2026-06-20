"""
EcoHotels Ratings — Professional Excel Beautifier (No Date Row)
Headers are BLACK & BOLD — missing headers get default names
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule

# ═══════════════════════════════════════════════════════════════════════════
#  USER SETTINGS – CHANGE THESE PATHS
# ═══════════════════════════════════════════════════════════════════════════

INPUT_PATH  = r"C:\Users\Administrator\Desktop\RatingsProject\Results\Eco_Reviews_&_Ratings.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\RatingsProject\Final_Result\Eco_Reviews_&_Ratings.xlsx"

# ═══════════════════════════════════════════════════════════════════════════
#  THEME: PROFESSIONAL LIGHT HEADERS WITH BLACK BOLD TEXT
# ═══════════════════════════════════════════════════════════════════════════
000000
HEADER_FONT_COLOR = "2C3E50"          # BLACK text
DEFAULT_HEADER_BG = "FFFFFF"          # Light gray background (always readable)
ALT_ROW_BG        = "F8FFFC"
BORDER_COLOR      = "2C3E50"
NA_COLOR          = "7F8C8D"

# Platform colors are no longer used for headers (to keep black text readable),
# but you can keep them for other purposes if needed.
PLATFORM_COLORS = {
    "agoda":       "C9A227",
    "booking":     "1D4ED8",
    "goibibo":     "DC2626",
    "google":      "16A34A",
    "mmt":         "0284C7",
    "tripadvisor": "059669",
}

LOW_RATING_FILL   = "FFF1F2"
COLOR_SCALE_START = "DC2626"
COLOR_SCALE_MID   = "FBBF24"
COLOR_SCALE_END   = "10B981"
DATA_BAR_COLOR    = "0F766E"

# Font settings
HEADER_FONT = "Calibri"
HEADER_SIZE = 10
HEADER_BOLD = True

BODY_FONT = "Century"
BODY_SIZE = 10

ROW_HEIGHT = 22
THIN_BORDER_STYLE = "thin"
FREEZE_PANES = "A2"
AUTO_FILTER = True
GRIDLINES = False

# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def platform_of(header):
    if not header: return None
    h = str(header).lower()
    for plat in PLATFORM_COLORS:
        if plat in h:
            return plat
    return None

def is_rating_column(header):
    return header and "rating" in str(header).lower()

def is_reviews_column(header):
    return header and "reviews" in str(header).lower()

def safe_float(value):
    if value is None: return None
    if isinstance(value, (int, float)): return float(value)
    try:
        cleaned = str(value).strip().replace(',', '')
        return float(cleaned)
    except:
        return None

def safe_int(value):
    if value is None: return None
    if isinstance(value, (int, float)): return int(value)
    try:
        cleaned = str(value).strip().replace(',', '')
        return int(float(cleaned))
    except:
        return None

# ═══════════════════════════════════════════════════════════════════════════
#  MAIN BEAUTIFICATION (headers: black + bold, missing names filled)
# ═══════════════════════════════════════════════════════════════════════════

def beautify_professional(input_path, output_path):
    print(f"Loading: {input_path}")
    wb = load_workbook(input_path)
    ws = wb.active

    # Store original header values BEFORE inserting title row
    original_headers = []
    for col in range(1, ws.max_column + 1):
        original_headers.append(ws.cell(1, col).value)

    # Insert title row at the top
    ws.insert_rows(1, amount=1)
    num_cols = ws.max_column

    # Write title row (row 1)
    title_cell = ws.cell(1, 1)
    title_cell.value = "🏨 Reviews & Ratings Performance Dashboard"
    title_cell.font = Font(name="Imprint MT Shadow", size=24, bold=True, color="2C3E50")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.row_dimensions[1].height = 40

    # Restore headers to row 2
    header_row = 2
    for col_idx, header_value in enumerate(original_headers, start=1):
        # If header is empty or None, generate a default name (Column A, B, ...)
        if header_value is None or str(header_value).strip() == "":
            col_letter = get_column_letter(col_idx)
            header_value = f"Column {col_letter}"
        ws.cell(header_row, col_idx).value = header_value

    # Read headers for processing
    headers = [ws.cell(header_row, c).value or "" for c in range(1, num_cols + 1)]
    print(f"Headers found (blanks replaced): {headers}")

    # Convert data to numbers
    first_data_row = header_row + 1
    last_data_row = ws.max_row

    for col_idx, hdr in enumerate(headers, start=1):
        if is_rating_column(hdr):
            for row_idx in range(first_data_row, last_data_row + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and str(cell.value).strip().upper() != "N/A":
                    num = safe_float(cell.value)
                    if num is not None: cell.value = num
        elif is_reviews_column(hdr):
            for row_idx in range(first_data_row, last_data_row + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and str(cell.value).strip().upper() != "N/A":
                    num = safe_int(cell.value)
                    if num is not None: cell.value = num

    # Style header row (BLACK, BOLD) with a uniform light background
    ws.row_dimensions[header_row].height = 32
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx)
        # Use uniform light gray background for all headers (ensures black text is readable)
        bg = DEFAULT_HEADER_BG
        cell.font = Font(name=HEADER_FONT, bold=HEADER_BOLD, color=HEADER_FONT_COLOR, size=HEADER_SIZE)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    # Data rows
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)
    base_font = Font(name=BODY_FONT, size=BODY_SIZE, color="1A1A1A")
    name_font = Font(name=BODY_FONT, size=BODY_SIZE, bold=True, color="1A1A1A")
    na_font = Font(name=BODY_FONT, size=BODY_SIZE, color=NA_COLOR, italic=True)

    for row_idx in range(first_data_row, last_data_row + 1):
        ws.row_dimensions[row_idx].height = 18
        is_alt = (row_idx % 2 == 0)

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            is_na = (value is not None and str(value).strip().upper() == "N/A")

            if is_alt:
                cell.fill = alt_fill
            else:
                cell.fill = PatternFill(fill_type=None)

            if col_idx == 1:
                cell.font = name_font
            elif is_na:
                cell.value = "—"
                cell.font = na_font
            else:
                cell.font = base_font

            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            cell.border = thin_border()

            if not is_na and value is not None:
                if is_rating_column(headers[col_idx-1]):
                    cell.number_format = "0.0"
                elif is_reviews_column(headers[col_idx-1]):
                    cell.number_format = "#,##0"

    # Conditional formatting
    for col_idx, hdr in enumerate(headers, start=1):
        if is_rating_column(hdr):
            col_letter = get_column_letter(col_idx)
            range_str = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
            rule = ColorScaleRule(
                start_type='min', start_color=COLOR_SCALE_START,
                mid_type='percentile', mid_value=50, mid_color=COLOR_SCALE_MID,
                end_type='max', end_color=COLOR_SCALE_END
            )
            ws.conditional_formatting.add(range_str, rule)
            low_fill = PatternFill(start_color=LOW_RATING_FILL, end_color=LOW_RATING_FILL, fill_type="solid")
            ws.conditional_formatting.add(range_str,
                CellIsRule(operator='lessThan', formula=['6'], stopIfTrue=True, fill=low_fill))

    for col_idx, hdr in enumerate(headers, start=1):
        if is_reviews_column(hdr):
            col_letter = get_column_letter(col_idx)
            range_str = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
            rule = DataBarRule(start_type='min', end_type='max', color=DATA_BAR_COLOR, showValue=True)
            ws.conditional_formatting.add(range_str, rule)

    # Column widths
    for col_idx, hdr in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 32
        elif is_rating_column(hdr):
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 13

    # Freeze panes & auto-filter
    ws.freeze_panes = f"B{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{last_data_row}"
    ws.sheet_properties.tabColor = DEFAULT_HEADER_BG

    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend["A1"] = "📊 Report Legend"
    legend["A1"].font = Font(name="Calibri", bold=True, size=12, color="2C3E50")
    legend["A3"] = "Conditional Formatting:"
    legend["A4"] = "  • Color scale: Red → Yellow → Green (ratings)"
    legend["A5"] = "  • Light red fill: Rating < 6.0"
    legend["A6"] = "  • Purple data bars: Review count"
    legend["A8"] = "Styling: Alternating light gray rows, thin borders, dash for N/A"
    legend["A9"] = "Column headers: BLACK BOLD text on light gray background"
    legend.column_dimensions["A"].width = 55

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Saved to: {output_path}")

if __name__ == "__main__":
    beautify_professional(INPUT_PATH, OUTPUT_PATH)